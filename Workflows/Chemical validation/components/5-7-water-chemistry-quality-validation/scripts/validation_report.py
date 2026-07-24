# -*- coding: utf-8 -*-
# ------------------------------------------------------------
# WATER CHEMISTRY VALIDATION REPORT
# ------------------------------------------------------------
# This script generates a validation report for ICP-Forest water
# chemistry samples from the validated CSV files produced by the
# WaterChemistryValidation wrapper.
#
# It produces three outputs:
#   1. validation_report.pdf  - multi-page PDF with key indicators,
#      charts and a colour-coded validation table per sample.
#   2. Samples2Repeat.xlsx    - Excel file with only the samples that
#      failed validation (FINAL_VALIDATION == NO), highlighted in red
#      where values are missing.
#   3. allFinalData.xlsx      - Excel file with all samples regardless
#      of validation outcome, same column layout and highlighting.
#
# Charts included in the PDF:
#   1. Failed validations by month (NOREP original vs REP replacement)
#   2. Failed validation type breakdown
#   3. Failure rate heatmap by sampling typology and month
#   4. Failure rate heatmap by sampling typology and validation criterion
#   5. Validation outcome by sampling typology
#   6. Failure rate by SiteCode and SiteName
#   7. REP improvement rate by typology
#   8. Failure rate by non-repeated sample
#   9. Failure rate by repeated sample
#
# Input:  ZIP of validated CSVs  +  samplesInfo.xlsx
# Output: PDF report  +  two Excel files
# ------------------------------------------------------------

# ============================================================
# IMPORTS
# ============================================================
import os
import re
import zipfile
import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")   # non-interactive backend - required inside Docker
import matplotlib.pyplot as plt
from fpdf import FPDF, XPos, YPos
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------------
# CLI argument parsing (Tesseract wrapper convention)
# No user-configurable parameters: report layout is fixed.
# ------------------------------------------------------------
parser = argparse.ArgumentParser(description="Water Chemistry Validation Report wrapper")
args = parser.parse_args()

# ------------------------------------------------------------
# Paths - Tesseract mounts inputs at /mnt/inputs, outputs at /mnt/outputs
# ------------------------------------------------------------
input_zip_path     = os.environ.get("INPUT_ZIP_PATH", "/mnt/inputs/water_chemical_alldata_validated.zip")
input_samples_path = os.environ.get("INPUT_SAMPLES_PATH", "/mnt/inputs/samplesInfo.xlsx")
output_pdf_path    = os.environ.get("OUTPUT_PDF_PATH", "/mnt/outputs/validation_report.pdf")
output_repeat_path = os.environ.get("OUTPUT_REPEAT_PATH", "/mnt/outputs/Samples2Repeat.xlsx")
output_all_path    = os.environ.get("OUTPUT_ALL_PATH", "/mnt/outputs/All_Validated_Data.xlsx")

# Temporary directories
extract_dir = os.environ.get("EXTRACT_DIR", "/tmp/input/level2_validated")
charts_dir  = os.environ.get("CHARTS_DIR", "/tmp/charts")
os.makedirs(extract_dir, exist_ok=True)
os.makedirs(charts_dir,  exist_ok=True)

# Extract input ZIP
with zipfile.ZipFile(input_zip_path, "r") as z:
    z.extractall(extract_dir)
print(f"Input ZIP extracted to {extract_dir}")

# Load sampling typology
if os.path.exists(input_samples_path):
    sampling_ty = pd.read_excel(input_samples_path)
    print(f"Sampling typology loaded: {len(sampling_ty)} records")
else:
    sampling_ty = pd.DataFrame(columns=["SampleID", "SamplingTypology"])
    print("WARNING: samplesInfo.xlsx not found - typology fill will be skipped.")

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_id(value):
    """Normalise SampleID: uppercase, remove spaces/underscores/hyphens/dots."""
    if pd.isna(value):
        return value
    return re.sub(r"[ _\-,.]", "", str(value).upper().strip())


def add_typology_from_sampling_info(df, sampling_ty):
    """
    Ensures that replicate (REP) samples receive their SamplingTypology
    from the matching non-REP entry in samplesInfo.xlsx.
    The join is done on a cleaned base SampleID (without the REP suffix).
    """
    d  = df.copy()
    st = sampling_ty.copy()

    if "SampleID" not in d.columns or "SampleID" not in st.columns:
        return d

    # Determine which column holds the typology
    if "SamplingTypology" in st.columns:
        type_col = "SamplingTypology"
    elif "Type" in st.columns:
        type_col = "Type"
    else:
        return d

    st["CleanSampleID_base"] = (
        st["SampleID"].apply(clean_id).astype(str).str.replace(r"REP$", "", regex=True)
    )
    st = (
        st[["CleanSampleID_base", type_col]]
        .drop_duplicates(subset=["CleanSampleID_base"])
        .rename(columns={type_col: "_Type_fill"})
    )

    d["CleanSampleID_base"] = (
        d["SampleID"].apply(clean_id).astype(str).str.replace(r"REP$", "", regex=True)
    )
    d = d.merge(st, on="CleanSampleID_base", how="left")

    # Fill missing typology values
    if "SamplingTypology" in d.columns:
        d["SamplingTypology"] = d["SamplingTypology"].fillna(d["_Type_fill"]).infer_objects(copy=False)
    elif "Type" in d.columns:
        d["Type"] = d["Type"].fillna(d["_Type_fill"]).infer_objects(copy=False)
    else:
        d["SamplingTypology"] = d["_Type_fill"]

    return d.drop(columns=["CleanSampleID_base", "_Type_fill"])


def prepare_report_data(data):
    """
    Standardises column names for the report layer and normalises
    validation values (NaN/None/empty → pd.NA).

    Renamed columns:
      sC - sA - Org- QualityIonsBalance -> Ions balance Org
      QualityConductivity               -> Conductivity
      QualityOrgN                       -> OrgN
      QualityRatioNa/Cl                 -> Ratio Na/Cl
      FINAL_VALIDATION                  -> VAL
      SamplingTypology                  -> Type
    """
    d = data.copy()
    d.rename(columns={
        "sC - sA - Org- QualityIonsBalance": "Ions balance Org",
        "QualityConductivity":               "Conductivity",
        "QualityOrgN":                       "OrgN",
        "QualityRatioNa/Cl":                 "Ratio Na/Cl",
        "FINAL_VALIDATION":                  "VAL",
        "SamplingTypology":                  "Type",
    }, inplace=True)

    for col in ["SampleID", "year", "month", "Type",
                "Ions balance Org", "Conductivity", "OrgN", "Ratio Na/Cl", "VAL"]:
        if col not in d.columns:
            d[col] = pd.NA

    # Normalise VAL
    d["VAL"] = (
        d["VAL"].astype(str).str.upper().str.strip()
        .replace({"NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
    )

    # Normalise quality columns
    for col in ["Ions balance Org", "Conductivity", "OrgN", "Ratio Na/Cl"]:
        d[col] = (
            d[col].astype(str).str.strip()
            .replace({"nan": pd.NA, "None": pd.NA, "NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
        )

    return d


def build_replacement_comparison(data):
    """
    Creates two paired scenarios with the same set of base samples:

    1. NOREP original:
       Validation results from the original non-replicate samples.

    2. NOREP replaced by REP:
       Same base samples, but validation results are replaced by the
       corresponding REP results wherever a REP exists for the same
       SampleID base / SiteCode / year / month.

    This allows assessing whether replicate measurements improve or
    worsen validation outcomes without changing the sample count.
    """
    d    = prepare_report_data(data)
    vcols = ["Ions balance Org", "Conductivity", "OrgN", "Ratio Na/Cl", "VAL"]

    d["SampleID_clean"] = d["SampleID"].apply(clean_id)
    d["IsRep"]          = d["SampleID_clean"].astype(str).str.contains(r"REP$", na=False)
    d["SampleID_base"]  = d["SampleID_clean"].astype(str).str.replace(r"REP$", "", regex=True)

    match_cols = (
        ["SampleID_base", "SiteCode", "year", "month"]
        if "SiteCode" in d.columns
        else ["SampleID_base", "year", "month"]
    )

    no_rep = d[~d["IsRep"]].copy()
    rep    = d[d["IsRep"]].copy()

    # Scenario 1 - original NOREP results
    original               = no_rep.copy()
    original["Scenario"]   = "NOREP original"
    original["SourceUsed"] = "NOREP"

    # If no REP samples exist, return only the original scenario
    # (avoids a duplicate "NOREP replaced by REP" that is identical to original)
    if rep.empty:
        original["VAL"] = (
            original["VAL"].astype(str).str.upper().str.strip()
            .replace({"NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
        )
        return original

    # Scenario 2 - replace validation values with REP where available
    rep_vals = (
        rep[match_cols + vcols]
        .drop_duplicates(subset=match_cols, keep="first")
        .rename(columns={c: f"{c}_REP" for c in vcols})
    )
    replaced = no_rep.merge(rep_vals, on=match_cols, how="left")
    for col in vcols:
        replaced[col] = replaced[f"{col}_REP"].combine_first(replaced[col])
    replaced = replaced.drop(columns=[f"{c}_REP" for c in vcols])
    replaced["Scenario"]   = "NOREP replaced by REP"
    replaced["SourceUsed"] = np.where(replaced["VAL"].notna(), "REP", "NOREP")

    comparison = pd.concat([original, replaced], ignore_index=True)
    comparison["VAL"] = (
        comparison["VAL"].astype(str).str.upper().str.strip()
        .replace({"NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
    )
    return comparison


def build_key_indicators(data):
    """
    Computes a set of summary statistics for a scenario DataFrame:
    total records, unique samples, sites, validated periods,
    failed/passed/unknown record counts, and overall failure rate.
    """
    df = data.copy()
    df["VAL"] = (
        df["VAL"].astype(str).str.upper().str.strip()
        .replace({"NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
    )

    total        = len(df)
    unique_sites = df["SiteCode"].nunique(dropna=True) if "SiteCode" in df.columns else 0
    periods      = df[["year", "month"]].drop_duplicates().shape[0] if {"year", "month"} <= set(df.columns) else 0
    failed       = int((df["VAL"] == "NO").sum())
    passed       = int((df["VAL"] == "SI").sum())
    unknown      = int(df["VAL"].isna().sum())
    failure_rate = failed / total * 100 if total > 0 else 0

    return [
        ("Total records",     total),
        ("Sites",             unique_sites),
        ("Validated periods", periods),
        ("Failed records",    failed),
        ("Passed records",    passed),
        ("Failure rate",      f"{failure_rate:.1f}%"),
        ("Unknown VAL",       unknown),
    ]


# ============================================================
# CHART FUNCTIONS
# ============================================================

def create_empty_chart(output_path, title):
    """Placeholder chart when no data is available for a panel."""
    fig, ax = plt.subplots(figsize=(8, 4.8))
    ax.text(0.5, 0.5, "No data available", ha="center", va="center", fontsize=14)
    ax.set_title(title)
    ax.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close()


def _sort_period_index(df):
    """Try to sort a Period-indexed DataFrame chronologically."""
    try:
        df.index = pd.to_datetime(df.index, format="%Y-%m", errors="coerce")
        df = df.sort_index()
        df.index = df.index.strftime("%Y-%m")
    except Exception:
        pass
    return df


def create_validation_charts(data, charts_dir):
    """
    Generates the five report charts and saves them as PNG files.

    Charts 1–3 compare NOREP original vs NOREP replaced by REP.
    Charts 4–5 show individual sample outcomes (all / REP only).
    """
    comparison_df = build_replacement_comparison(data)
    actual_df     = prepare_report_data(data)

    # True only when the dataset actually contains REP samples
    has_rep = actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False).any()

    SCENARIOS = ["NOREP original", "NOREP replaced by REP"]
    COLORS_BAR  = ["#AFCBFF", "#FFD6A5"]
    COLORS_LINE = ["#5B8DEF", "#F4A261"]

    for df in [comparison_df, actual_df]:
        df["year"]  = pd.to_numeric(df["year"],  errors="coerce")
        df["month"] = pd.to_numeric(df["month"], errors="coerce")
        df["Period"] = np.where(
            df["year"].notna() & df["month"].notna(),
            df["year"].astype("Int64").astype(str) + "-" +
            df["month"].astype("Int64").astype(str).str.zfill(2),
            "Unknown",
        )
        df["VAL"] = (
            df["VAL"].astype(str).str.upper().str.strip()
            .replace({"NAN": pd.NA, "NONE": pd.NA, "": pd.NA})
        )

    chart_paths = []

    # ----------------------------------------------------------
    # Chart 1 - Failed validations by month
    # ----------------------------------------------------------
    path1 = os.path.join(charts_dir, "chart_fail_by_month.png")
    fail_by_month = (
        comparison_df.assign(IsFail=comparison_df["VAL"] == "NO")
        .groupby(["Period", "Scenario"])["IsFail"].sum()
        .unstack(fill_value=0)
    )
    fail_by_month = _sort_period_index(fail_by_month)

    if not fail_by_month.empty:
        fig, ax = plt.subplots(figsize=(8, 4.8))
        cols = [c for c in SCENARIOS if c in fail_by_month.columns]
        fail_by_month[cols].plot(kind="bar", ax=ax, color=COLORS_BAR[:len(cols)])
        ax.set_title("Failed validations: original vs REP replacement" if has_rep else "Failed validations by month")
        ax.set_xlabel("Month")
        ax.set_ylabel("Failed records")
        ax.tick_params(axis="x", rotation=45)
        ax.legend(title="Scenario")
        plt.tight_layout()
        plt.savefig(path1, dpi=200, bbox_inches="tight")
        plt.close()
    else:
        create_empty_chart(path1, "Failed validations by month")
    chart_paths.append(path1)

    # ----------------------------------------------------------
    # Chart 2 - Failure type breakdown
    # ----------------------------------------------------------
    path2 = os.path.join(charts_dir, "chart_failure_type.png")
    validation_rules = {
        "Ions balance Org": ("Ions balance Org", "NO"),
        "Conductivity":     ("Conductivity",     "NO"),
        "OrgN":             ("OrgN",             "NO TN"),
        "Ratio Na/Cl":      ("Ratio Na/Cl",      "NO"),
    }
    failure_type = pd.DataFrame(index=validation_rules.keys())
    for scenario in SCENARIOS:
        sdf = comparison_df[comparison_df["Scenario"] == scenario]
        counts = {}
        for rule, (col, val) in validation_rules.items():
            counts[rule] = (sdf[col].astype(str).str.upper().str.strip() == val).sum() if col in sdf.columns else 0
        failure_type[scenario] = pd.Series(counts)
    failure_type = failure_type.loc[failure_type.sum(axis=1).sort_values(ascending=False).index]

    if not failure_type.empty:
        fig, ax = plt.subplots(figsize=(8, 4.8))
        cols = [c for c in SCENARIOS if c in failure_type.columns]
        failure_type[cols].plot(kind="bar", ax=ax, color=COLORS_BAR[:len(cols)])
        ax.set_title("Failed validation type: original vs REP replacement" if has_rep else "Failed validation type")
        ax.set_xlabel("Validation type")
        ax.set_ylabel("Number of failures")
        ax.tick_params(axis="x", rotation=20)
        ax.legend(title="Scenario")
        plt.tight_layout()
        plt.savefig(path2, dpi=200, bbox_inches="tight")
        plt.close()
    else:
        create_empty_chart(path2, "Failed validation type")
    chart_paths.append(path2)

    # ----------------------------------------------------------
    # Chart 3 - Heatmap: failure rate by typology x month
    # Shows which typology fails more in which month, helping
    # identify seasonal patterns per sampling type.
    # Only rendered when typology data exists.
    # ----------------------------------------------------------
    path3 = os.path.join(charts_dir, "chart_heatmap_typology_month.png")

    type_col_pre = "Type" if "Type" in actual_df.columns else None
    has_typo_pre = (
        type_col_pre is not None and
        actual_df[type_col_pre].notna().any() and
        (actual_df[type_col_pre].astype(str).str.strip() != "").any()
    )

    if has_typo_pre:
        hm_df = actual_df[actual_df[type_col_pre].notna()].copy()
        hm_df[type_col_pre] = hm_df[type_col_pre].astype(str).str.strip()
        hm_df["IsFail"] = (hm_df["VAL"] == "NO").astype(int)

        # Pivot: rows = typology, cols = period, values = failure rate %
        hm_total  = hm_df.groupby([type_col_pre, "Period"]).size().unstack(fill_value=0)
        hm_failed = hm_df.groupby([type_col_pre, "Period"])["IsFail"].sum().unstack(fill_value=0)
        hm_rate   = (hm_failed / hm_total.replace(0, np.nan) * 100).fillna(0)

        # Sort columns chronologically
        try:
            sorted_cols = sorted(hm_rate.columns, key=lambda x: pd.to_datetime(x, format="%Y-%m", errors="coerce"))
            hm_rate = hm_rate[sorted_cols]
        except Exception:
            pass

        # Sort rows by total failure rate descending
        hm_rate = hm_rate.loc[hm_rate.mean(axis=1).sort_values(ascending=False).index]

        fig, ax = plt.subplots(figsize=(max(8, len(hm_rate.columns) * 0.7), max(4, len(hm_rate.index) * 0.55)))
        im = ax.imshow(hm_rate.values, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=100)

        ax.set_xticks(range(len(hm_rate.columns)))
        ax.set_xticklabels(hm_rate.columns, rotation=45, ha="right", fontsize=7)
        ax.set_yticks(range(len(hm_rate.index)))
        ax.set_yticklabels(hm_rate.index, fontsize=7)
        ax.set_title("Failure rate (%) by typology and month")
        ax.set_xlabel("Month")
        ax.set_ylabel("Sampling typology")

        # Add % values inside cells
        for i in range(len(hm_rate.index)):
            for j in range(len(hm_rate.columns)):
                val = hm_rate.values[i, j]
                ax.text(j, i, f"{val:.0f}%", ha="center", va="center",
                        fontsize=6, color="black" if val < 60 else "white")

        plt.colorbar(im, ax=ax, label="Failure rate (%)", fraction=0.03)
        plt.tight_layout()
        plt.savefig(path3, dpi=200, bbox_inches="tight")
        plt.close()
    else:
        create_empty_chart(path3, "Failure rate by typology and month (no typology data)")
    chart_paths.append(path3)

    # ----------------------------------------------------------
    # Chart 4 - Heatmap: failure rate by typology x validation criterion
    # Rows = sampling typology; columns = validation criteria.
    # This complements the month heatmap by showing which validation
    # rule is driving failures within each type of water sample.
    # ----------------------------------------------------------
    path4 = os.path.join(charts_dir, "chart_heatmap_typology_validation.png")

    if has_typo_pre:
        hv_df = actual_df[actual_df[type_col_pre].notna()].copy()
        hv_df[type_col_pre] = hv_df[type_col_pre].astype(str).str.strip()

        validation_rules_hm = {
            "Ions balance Org": ("Ions balance Org", "NO"),
            "Ratio Na/Cl":      ("Ratio Na/Cl",      "NO"),
            "Conductivity":     ("Conductivity",     "NO"),
            "OrgN":             ("OrgN",             "NO TN"),
        }

        rows = []
        for typology, g in hv_df.groupby(type_col_pre):
            total = len(g)
            row = {type_col_pre: typology}
            for label, (col, fail_value) in validation_rules_hm.items():
                if col in g.columns and total > 0:
                    failures = (g[col].astype(str).str.upper().str.strip() == fail_value).sum()
                    row[label] = failures / total * 100
                else:
                    row[label] = np.nan
            rows.append(row)

        hv_rate = pd.DataFrame(rows).set_index(type_col_pre)
        hv_rate = hv_rate[validation_rules_hm.keys()]
        hv_rate = hv_rate.fillna(0)
        hv_rate = hv_rate.loc[hv_rate.mean(axis=1).sort_values(ascending=False).index]

        if not hv_rate.empty:
            fig, ax = plt.subplots(figsize=(8, max(4, len(hv_rate.index) * 0.55)))
            im = ax.imshow(hv_rate.values, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=100)

            ax.set_xticks(range(len(hv_rate.columns)))
            ax.set_xticklabels(hv_rate.columns, rotation=25, ha="right", fontsize=8)
            ax.set_yticks(range(len(hv_rate.index)))
            ax.set_yticklabels(hv_rate.index, fontsize=7)
            ax.set_title("Failure rate (%) by typology and validation criterion")
            ax.set_xlabel("Validation criterion")
            ax.set_ylabel("Sampling typology")

            for i in range(len(hv_rate.index)):
                for j in range(len(hv_rate.columns)):
                    val = hv_rate.values[i, j]
                    ax.text(j, i, f"{val:.0f}%", ha="center", va="center",
                            fontsize=7, color="black" if val < 60 else "white")

            plt.colorbar(im, ax=ax, label="Failure rate (%)", fraction=0.04)
            plt.tight_layout()
            plt.savefig(path4, dpi=200, bbox_inches="tight")
            plt.close()
        else:
            create_empty_chart(path4, "Failure rate by typology and validation criterion")
    else:
        create_empty_chart(path4, "Failure rate by typology and validation criterion (no typology data)")
    chart_paths.append(path4)

    # ----------------------------------------------------------
    # Chart 5 - Failures by SamplingTypology (only when typology info exists)
    # ----------------------------------------------------------
    path4 = os.path.join(charts_dir, "chart_fail_by_typology.png")

    # Detect which column holds the typology after prepare_report_data renames it
    type_col = "Type" if "Type" in actual_df.columns else None
    has_typology = (
        type_col is not None and
        actual_df[type_col].notna().any() and
        (actual_df[type_col].astype(str).str.strip() != "").any()
    )

    if has_typology:
        typo_df = actual_df[actual_df[type_col].notna()].copy()
        typo_df[type_col] = typo_df[type_col].astype(str).str.strip()

        # Separate NOREP and REP rows
        is_rep  = typo_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
        norep_df = typo_df[~is_rep]
        rep_df   = typo_df[is_rep]

        def _typo_counts(df, type_col):
            return (
                df.groupby(type_col)["VAL"]
                .agg(
                    Failures=lambda x: (x == "NO").sum(),
                    Passes=lambda x:   (x == "SI").sum(),
                )
                .reset_index()
            )

        norep_summary = _typo_counts(norep_df, type_col)
        # Sort by NOREP failures descending
        norep_summary = norep_summary.sort_values("Failures", ascending=False).reset_index(drop=True)
        typology_order = norep_summary[type_col].tolist()

        fig, ax = plt.subplots(figsize=(10, 4.8))
        x     = np.arange(len(typology_order))
        width = 0.35 if has_rep else 0.5

        # NOREP stacked bars
        ax.bar(x - (width / 2 if has_rep else 0),
               norep_summary["Failures"], width=width,
               label="NOREP Failures", color="#F4A6A6")
        ax.bar(x - (width / 2 if has_rep else 0),
               norep_summary["Passes"],  width=width,
               bottom=norep_summary["Failures"],
               label="NOREP Passes", color="#A8D5BA")

        # REP stacked bars (only when REP samples exist)
        if has_rep and not rep_df.empty:
            rep_summary = _typo_counts(rep_df, type_col)
            # Align to the same typology order, fill missing with 0
            rep_summary = (
                pd.DataFrame({type_col: typology_order})
                .merge(rep_summary, on=type_col, how="left")
                .fillna(0)
            )
            ax.bar(x + width / 2,
                   rep_summary["Failures"], width=width,
                   label="REP Failures", color="#E07070")
            ax.bar(x + width / 2,
                   rep_summary["Passes"],  width=width,
                   bottom=rep_summary["Failures"],
                   label="REP Passes", color="#6DB88A")

        ax.set_xticks(x)
        ax.set_xticklabels(typology_order, rotation=30, ha="right", fontsize=8)
        ax.set_title("Validation outcome by sampling typology")
        ax.set_xlabel("Sampling typology")
        ax.set_ylabel("Number of records")
        ax.legend(fontsize=7)
        plt.tight_layout()
        plt.savefig(path4, dpi=200, bbox_inches="tight")
        plt.close()
    else:
        create_empty_chart(path4, "Validation failures by sampling typology (no typology data)")

    chart_paths.append(path4)

    # ----------------------------------------------------------
    # Chart 6 - Top SiteCodes by failure rate (NOREP vs REP)
    # Shows which sites fail most. With 14 sites a horizontal bar
    # sorted by NOREP failure rate is readable. REP bar shown when
    # REP samples exist, allowing direct site-level comparison.
    # ----------------------------------------------------------
    path5 = os.path.join(charts_dir, "chart_fail_by_sitecode.png")

    if "SiteCode" in actual_df.columns:
        is_rep_site  = actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
        norep_site   = actual_df[~is_rep_site].copy()
        rep_site     = actual_df[is_rep_site].copy()

        def _site_fail_rate(df):
            g = df.groupby("SiteCode")["VAL"].agg(
                Failures=lambda x: (x == "NO").sum(),
                Total=lambda x: len(x),
            ).reset_index()
            g["FailRate"] = g["Failures"] / g["Total"].replace(0, np.nan) * 100
            return g

        def _site_label_map(df):
            if "SiteName" not in df.columns:
                return {}
            site_meta = (
                df[["SiteCode", "SiteName"]]
                .dropna(subset=["SiteCode"])
                .drop_duplicates(subset=["SiteCode"], keep="first")
            )
            labels = {}
            for _, r in site_meta.iterrows():
                code = str(r["SiteCode"]).strip()
                name = "" if pd.isna(r["SiteName"]) else str(r["SiteName"]).strip()
                labels[r["SiteCode"]] = f"{code} - {name}" if name else code
            return labels

        site_labels = _site_label_map(actual_df)

        norep_site_g = _site_fail_rate(norep_site).sort_values("FailRate", ascending=True)
        site_order   = norep_site_g["SiteCode"].tolist()
        y_labels     = [site_labels.get(code, str(code)) for code in site_order]

        fig, ax = plt.subplots(figsize=(9.5, max(4.8, len(site_order) * 0.38)))
        y       = np.arange(len(site_order))
        height  = 0.35 if has_rep else 0.6

        ax.barh(y - (height / 2 if has_rep else 0),
                norep_site_g["FailRate"], height=height,
                label="NOREP", color="#AFCBFF")

        if has_rep and not rep_site.empty:
            rep_site_g = _site_fail_rate(rep_site)
            rep_site_g = (
                pd.DataFrame({"SiteCode": site_order})
                .merge(rep_site_g, on="SiteCode", how="left")
                .fillna(0)
            )
            ax.barh(y + height / 2, rep_site_g["FailRate"], height=height,
                    label="REP", color="#FFD6A5")

        ax.set_yticks(y)
        ax.set_yticklabels(y_labels, fontsize=7)
        ax.set_xlabel("Failure rate (%)")
        ax.set_title("Failure rate by SiteCode and SiteName" + (" (NOREP vs REP)" if has_rep else ""))
        ax.set_xlim(0, 105)
        ax.legend(fontsize=8)
        plt.tight_layout()
        plt.savefig(path5, dpi=200, bbox_inches="tight")
        plt.close()
    else:
        create_empty_chart(path5, "Failure rate by SiteCode (no SiteCode data)")
    chart_paths.append(path5)

    # ----------------------------------------------------------
    # Chart 7 - REP improvement rate by typology
    # Of the samples that FAILED as NOREP, shows per typology:
    #   - % where REP improved (NO -> SI)
    #   - % where REP did not improve (NO -> NO)
    #   - % with no REP available
    # Only rendered when both REP samples and typology data exist.
    # ----------------------------------------------------------
    path6 = os.path.join(charts_dir, "chart_rep_improvement.png")

    if has_rep and has_typo_pre:
        # Build matched NOREP/REP pairs
        norep_imp = actual_df[
            ~actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
            & (actual_df["VAL"] == "NO")
        ].copy()

        norep_imp["SampleID_clean"] = norep_imp["SampleID"].apply(clean_id)
        norep_imp["SampleID_base"]  = norep_imp["SampleID_clean"]

        rep_imp = actual_df[
            actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
        ].copy()
        rep_imp["SampleID_base"] = (
            rep_imp["SampleID"].apply(clean_id).astype(str).str.replace(r"REP$", "", regex=True)
        )

        match_cols_imp = ["SampleID_base", "year", "month"]
        if "SiteCode" in norep_imp.columns:
            match_cols_imp = ["SampleID_base", "SiteCode", "year", "month"]

        rep_vals_imp = (
            rep_imp[match_cols_imp + ["VAL"]]
            .drop_duplicates(subset=match_cols_imp)
            .rename(columns={"VAL": "VAL_REP"})
        )

        merged_imp = norep_imp.merge(rep_vals_imp, on=match_cols_imp, how="left")
        merged_imp[type_col_pre] = merged_imp[type_col_pre].astype(str).str.strip()

        # Classify each row
        merged_imp["Outcome"] = np.select(
            [
                merged_imp["VAL_REP"] == "SI",
                merged_imp["VAL_REP"] == "NO",
            ],
            ["Improved (NO -> SI)", "Not improved (NO -> NO)"],
            default="No REP available"
        )

        imp_summary = (
            merged_imp.groupby([type_col_pre, "Outcome"]).size()
            .unstack(fill_value=0)
            .reset_index()
        )

        # Ensure all three outcome columns exist
        for col in ["Improved (NO -> SI)", "Not improved (NO -> NO)", "No REP available"]:
            if col not in imp_summary.columns:
                imp_summary[col] = 0

        # Sort by improvement rate descending
        imp_total = imp_summary[["Improved (NO -> SI)", "Not improved (NO -> NO)", "No REP available"]].sum(axis=1)
        imp_summary["ImpRate"] = imp_summary["Improved (NO -> SI)"] / imp_total.replace(0, np.nan) * 100
        imp_summary = imp_summary.sort_values("ImpRate", ascending=False).reset_index(drop=True)

        if not imp_summary.empty:
            fig, ax = plt.subplots(figsize=(8, 4.8))
            x     = np.arange(len(imp_summary))
            width = 0.25

            ax.bar(x - width, imp_summary["Improved (NO -> SI)"],
                   width=width, label="Improved (NO -> SI)",     color="#A8D5BA")
            ax.bar(x,         imp_summary["Not improved (NO -> NO)"],
                   width=width, label="Not improved (NO -> NO)", color="#F4A6A6")
            ax.bar(x + width, imp_summary["No REP available"],
                   width=width, label="No REP available",        color="#D0D0D0")

            ax.set_xticks(x)
            ax.set_xticklabels(imp_summary[type_col_pre], rotation=30, ha="right", fontsize=8)
            ax.set_title("REP improvement rate by typology (failed NOREP samples only)")
            ax.set_xlabel("Sampling typology")
            ax.set_ylabel("Number of samples")
            ax.legend(fontsize=7)
            plt.tight_layout()
            plt.savefig(path6, dpi=200, bbox_inches="tight")
            plt.close()
        else:
            create_empty_chart(path6, "REP improvement rate by typology")
    else:
        msg = "REP improvement by typology (no REP data)" if not has_rep else "REP improvement by typology (no typology data)"
        create_empty_chart(path6, msg)
    chart_paths.append(path6)

    # ----------------------------------------------------------
    # ----------------------------------------------------------
    # Charts 8 & 9 - Failure rate ranking by SampleID
    # One bar per sample, sorted by failure rate % descending.
    # Bar colour is a red-green gradient based on failure rate.
    # Total records shown as annotation at end of each bar.
    # Height scales dynamically with number of samples.
    # ----------------------------------------------------------

    def _failure_rate_chart(summary, title, output_path):
        if summary.empty:
            create_empty_chart(output_path, title)
            return
        df = summary.copy()
        df["Total"]    = df["Failures"] + df["Passes"]
        df["FailRate"] = df["Failures"] / df["Total"].replace(0, np.nan) * 100
        df = df.sort_values("FailRate", ascending=True).reset_index(drop=True)

        # Build y-axis label: "SAMPLEID  [Typology]" when typology is available
        if "Type" in df.columns:
            df["Label"] = df.apply(
                lambda r: f"{r['SampleID']}  [{r['Type']}]"
                if pd.notna(r["Type"]) and str(r["Type"]).strip() not in ("", "nan", "None")
                else str(r["SampleID"]),
                axis=1
            )
        else:
            df["Label"] = df["SampleID"]

        n     = len(df)
        fig_h = max(5, n * 0.32)
        y_fs  = 8 if n <= 40 else (6 if n <= 80 else 5)

        fig, ax = plt.subplots(figsize=(12.5, fig_h))

        cmap   = plt.cm.RdYlGn_r
        colors = [cmap(rate / 100) for rate in df["FailRate"].fillna(0)]
        ax.barh(df["Label"], df["FailRate"], color=colors, edgecolor="white", linewidth=0.3)

        for i, (rate, total) in enumerate(zip(df["FailRate"].fillna(0), df["Total"])):
            ax.text(rate + 0.5, i, f"n={int(total)}", va="center", fontsize=y_fs - 1, color="#444444")

        ax.set_xlim(0, 115)
        ax.set_xlabel("Failure rate (%)", fontsize=9)
        ax.set_ylabel("SampleID", fontsize=9)
        ax.set_title(title, fontsize=10)
        ax.tick_params(axis="y", labelsize=y_fs)
        ax.tick_params(axis="x", labelsize=8)
        ax.invert_yaxis()

        sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(vmin=0, vmax=100))
        sm.set_array([])
        plt.colorbar(sm, ax=ax, label="Failure rate (%)", fraction=0.015, pad=0.01)

        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close()

    path7 = os.path.join(charts_dir, "chart_validation_by_sample.png")
    no_rep_df = actual_df[
        ~actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
    ].copy()
    # Include Type in groupby so the chart label can show typology
    grp_cols_norep = ["SampleID", "Type"] if "Type" in no_rep_df.columns else ["SampleID"]
    sample_summary = no_rep_df.groupby(grp_cols_norep)["VAL"].agg(
        Failures=lambda x: (x == "NO").sum(),
        Passes=lambda x:   (x == "SI").sum(),
    ).reset_index()
    _failure_rate_chart(sample_summary, "Failure rate by sample (non-repeated)", path7)
    chart_paths.append(path7)

    path8 = os.path.join(charts_dir, "chart_validation_by_repeated_sample.png")
    rep_df_s = actual_df[
        actual_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False)
    ].copy()
    grp_cols_rep = ["SampleID", "Type"] if "Type" in rep_df_s.columns else ["SampleID"]
    rep_summary = rep_df_s.groupby(grp_cols_rep)["VAL"].agg(
        Failures=lambda x: (x == "NO").sum(),
        Passes=lambda x:   (x == "SI").sum(),
    ).reset_index()
    _failure_rate_chart(rep_summary, "Failure rate by sample (repeated)", path8)
    chart_paths.append(path8)

    return chart_paths


# ============================================================
# PDF GENERATION
# ============================================================

def _draw_key_indicators_table(pdf, indicators, title):
    """Renders a standard two-column key-indicators table."""
    pdf.set_font("Helvetica", style="B", size=10)
    pdf.cell(0, 6, title, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(1)

    col_widths = [125, 35]
    pdf.set_font("Helvetica", style="B", size=8)
    pdf.set_fill_color(200, 220, 255)
    for header, w in zip(["Metric", "Value"], col_widths):
        pdf.cell(w, 5, header, border=1, align="C", fill=True)
    pdf.ln(5)

    pdf.set_font("Helvetica", size=8)
    for metric, value in indicators:
        pdf.cell(col_widths[0], 5, str(metric), border=1)
        pdf.cell(col_widths[1], 5, str(value), border=1, align="C")
        pdf.ln(5)
    pdf.ln(2)


def _draw_quality_notes(pdf):
    """Adds interpretation notes for the quality criteria."""
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", style="B", size=10)
    pdf.cell(0, 6, "Notes on the interpretation of quality criteria for ion balance and conductivity",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(1)

    notes = [
        "a) Quality criteria for conductivity must always be satisfied (OK) for each type of sample: wet only, bulk open field, throughfall, stemflow, soil water and runoff.",
        "b) Quality criteria for ion balance should be satisfied (OK) for open field samples (wet only and bulk) and runoff with low organic carbon concentrations. The ion balance criterion is not considered for throughfall and stemflow samples because of the presence of organic anions.",
        "c) Quality criteria for ON must always be satisfied (OK) for each type of sample: open field, throughfall, stemflow, soil water and runoff.",
        "d) Quality criteria for Na/Cl ratio (marine ratio = 0.86, accepted range 0.5-1.5) should be satisfied for each type of sample, excluding soil water and runoff samples.",
    ]

    pdf.set_font("Helvetica", size=7.5)
    for note in notes:
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 4.2, note, align="L", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_x(pdf.l_margin)
    pdf.ln(2)


def _draw_subprogrammes_table(pdf):
    """Adds the ICP Forests water subprogrammes table."""
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", style="B", size=10)
    pdf.cell(0, 6, "1.2 Water subprogrammes", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(1)

    headers = ["Subprogram", "Code", "Description"]
    col_widths = [48, 18, 124]
    rows = [
        ("Precipitation Chemistry", "PC", "Bulk deposition in open areas - quantifies wet deposition inputs"),
        ("Throughfall", "TF", "Water passing through the forest canopy - reflects canopy interactions"),
        ("Stemflow", "SF", "Water flowing down tree stems - localised flux insights"),
        ("Soil Water", "SW", "Soil solution chemistry - acidification and nitrogen dynamics"),
        ("Runoff Water", "RW", "Catchment outflows - estimates element export via hydrology"),
    ]

    pdf.set_font("Helvetica", style="B", size=7.5)
    pdf.set_fill_color(200, 220, 255)
    for header, w in zip(headers, col_widths):
        pdf.cell(w, 5, header, border=1, align="C", fill=True)
    pdf.ln(5)

    pdf.set_font("Helvetica", size=7.2)
    for row in rows:
        for value, w in zip(row, col_widths):
            pdf.cell(w, 6, value, border=1)
        pdf.ln(6)
    pdf.ln(3)

def _place_image_fit(pdf, image_path, x, y, box_w, box_h, align="center"):
    """Places an image inside a bounding box without distorting it."""
    try:
        with Image.open(image_path) as img:
            img_w, img_h = img.size
    except Exception:
        pdf.image(image_path, x=x, y=y, w=box_w, h=box_h)
        return

    if img_w <= 0 or img_h <= 0:
        pdf.image(image_path, x=x, y=y, w=box_w, h=box_h)
        return

    scale = min(box_w / img_w, box_h / img_h)
    draw_w = img_w * scale
    draw_h = img_h * scale

    if align == "left":
        draw_x = x
    elif align == "right":
        draw_x = x + (box_w - draw_w)
    else:
        draw_x = x + (box_w - draw_w) / 2

    draw_y = y + (box_h - draw_h) / 2
    pdf.image(image_path, x=draw_x, y=draw_y, w=draw_w, h=draw_h)


def _draw_table_header(pdf, columns, col_widths):
    """Renders the validation table header, supporting two-line column names."""
    two_line = {"Ions balance Org": ("Ions", "balance Org"), "Ratio Na/Cl": ("Ratio", "Na/Cl")}
    h = 5
    pdf.set_font("Helvetica", style="B", size=9)
    pdf.set_fill_color(200, 220, 255)
    y0 = pdf.get_y()

    for col, w in zip(columns, col_widths):
        x = pdf.get_x()
        if col in two_line:
            t1, t2 = two_line[col]
            pdf.cell(w, 2 * h, "", border=1, fill=True)
            pdf.set_xy(x, y0);       pdf.cell(w, h, t1, border=0, align="C")
            pdf.set_xy(x, y0 + h);  pdf.cell(w, h, t2, border=0, align="C")
            pdf.set_xy(x + w, y0)
        else:
            pdf.cell(w, 2 * h, col, border=1, align="C", fill=True)
    pdf.ln(2 * h)


def generate_pdf_report(df_final, sampling_ty, chart_paths, output_path):
    """
    Assembles the multi-page PDF report:
      Page 1: title, interpretation notes, water subprogrammes and key indicators
      Page 2: graphical summary heatmaps and overview charts
      Page 3: additional graphical summary charts
      Page 4: validation by non-repeated sample
      Page 5: validation by repeated sample (if REP exist)
      Page N: colour-coded detailed validation table
    """
    data = prepare_report_data(df_final)

    comparison   = build_replacement_comparison(data)
    ind_norep    = build_key_indicators(comparison[comparison["Scenario"] == "NOREP original"])
    ind_rep      = build_key_indicators(comparison[comparison["Scenario"] == "NOREP replaced by REP"])
    has_rep      = data["SampleID"].astype(str).str.upper().str.contains("REP", na=False).any()

    # Build sorted SampleID order from samplesInfo
    samplesIDOrden = [
        clean_id(x) for x in sampling_ty["SampleID"].tolist()
        if pd.notna(x) and str(x).strip() != ""
    ]
    combined_order = []
    for x in samplesIDOrden:
        combined_order.append(x)
        combined_order.append(x + "REP")

    # Prepare table data
    table_df = data[[
        "SampleID", "year", "month", "Type",
        "Ions balance Org", "Conductivity", "OrgN", "Ratio Na/Cl", "VAL",
    ]].copy()
    table_df["REP"] = np.where(
        table_df["SampleID"].astype(str).str.upper().str.contains("REP", na=False),
        "Rep", "No Rep"
    )
    table_df["year"]  = pd.to_numeric(table_df["year"],  errors="coerce")
    table_df["month"] = pd.to_numeric(table_df["month"], errors="coerce")
    table_df["_clean"] = table_df["SampleID"].apply(clean_id)
    table_df["_order"] = pd.Categorical(table_df["_clean"], categories=combined_order, ordered=True)
    table_df = table_df.sort_values(by=["year", "month", "_order"]).drop(columns=["_clean", "_order"])
    table_df["SampleID"] = table_df["SampleID"].apply(clean_id)
    table_df["month"] = table_df["month"].astype("Int64").astype(str)

    # Remove empty/NA SampleID rows
    table_df = table_df[
        table_df["SampleID"].notna() &
        (table_df["SampleID"].astype(str).str.strip() != "") &
        (table_df["SampleID"].astype(str).str.upper().str.strip() != "NA")
    ].reset_index(drop=True)

    TABLE_COLS   = ["SampleID", "month", "Type", "Ions balance Org",
                    "Conductivity", "OrgN", "Ratio Na/Cl", "REP", "VAL"]
    TABLE_WIDTHS = [30, 13, 23, 21, 21, 21, 21, 18, 15]

    FILL_COLORS = {
        "NO":     (255,   0,   0),
        "NO TN":  (255,   0,   0),
        "OK":     (  0, 255,   0),
        "SI":     (  0, 255,   0),
        "REP":    (255, 230, 180),
        "NO REP": (255, 255, 255),
    }

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    # ---- Page 1 - introduction, interpretation notes and key indicators ----
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=16)
    pdf.cell(200, 10, "WATER CHEMICAL VALIDATION REPORT FOR ICP SAMPLES", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", size=9)
    pdf.multi_cell(0, 5,
        "This report summarises the validation results of water sample data. "
        "It compares the original non-repeated sample validation results against "
        "a scenario where the validation results are replaced by their replicate (REP) "
        "counterparts when a REP exists. The comparison keeps the same number of samples "
        "in both scenarios. Separate charts are also provided for all samples and "
        "replicate-only samples.",
        align="J",
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT
    )
    pdf.set_x(pdf.l_margin)
    pdf.ln(2)

    _draw_quality_notes(pdf)
    _draw_subprogrammes_table(pdf)

    _draw_key_indicators_table(pdf, ind_norep, "Key Indicators - NOREP Original")
    if has_rep:
        _draw_key_indicators_table(pdf, ind_rep, "Key Indicators - NOREP Replaced by REP")

    # ---- Page 2 - graphical summary: overview and heatmaps ----
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=14)
    pdf.cell(0, 10, "Graphical Summary - Overview and Heatmaps", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(2)

    W = 88
    H = 76
    G = 10
    y0 = pdf.get_y()
    y1 = y0 + H + G

    _place_image_fit(pdf, chart_paths[0], x=10,  y=y0, box_w=W, box_h=H)   # failed validations by month
    _place_image_fit(pdf, chart_paths[1], x=110, y=y0, box_w=W, box_h=H)   # failure type breakdown
    _place_image_fit(pdf, chart_paths[2], x=10,  y=y1, box_w=W, box_h=H)   # heatmap typology x month
    _place_image_fit(pdf, chart_paths[3], x=110, y=y1, box_w=W, box_h=H)   # heatmap typology x validation criterion

    # ---- Page 3 - graphical summary: typology, SiteCode/SiteName and REP improvement ----
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=14)
    pdf.cell(0, 10, "Graphical Summary - Typology, Site and REP Effects", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(2)

    W2 = 88
    H2 = 72
    y0 = pdf.get_y()
    y1 = y0 + H2 + 12

    _place_image_fit(pdf, chart_paths[4], x=10,  y=y0, box_w=W2, box_h=H2)   # validation outcome by typology
    _place_image_fit(pdf, chart_paths[5], x=110, y=y0, box_w=W2, box_h=H2)   # failure rate by SiteCode + SiteName
    _place_image_fit(pdf, chart_paths[6], x=10,  y=y1, box_w=188, box_h=H2)  # REP improvement by typology

    # ---- Page 4 - outcome by non-repeated sample ----
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=14)
    pdf.cell(0, 10, "Validation Outcome by Non-Repeated Samples", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(3)
    _place_image_fit(pdf, chart_paths[7], x=10, y=pdf.get_y(), box_w=190, box_h=240)

    # ---- Page 5 - outcome by repeated sample (only if REP exist) ----
    if has_rep:
        pdf.add_page()
        pdf.set_font("Helvetica", style="B", size=14)
        pdf.cell(0, 10, "Validation Outcome by Repeated Samples", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
        pdf.ln(3)
        _place_image_fit(pdf, chart_paths[8], x=10, y=pdf.get_y(), box_w=190, box_h=240)

    # ---- Detailed table ----
    pdf.add_page()
    pdf.set_font("Helvetica", style="B", size=14)
    pdf.cell(0, 10, "Validation Table by Sample", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
    pdf.ln(5)
    _draw_table_header(pdf, TABLE_COLS, TABLE_WIDTHS)

    pdf.set_font("Helvetica", size=8)
    for _, row in table_df.iterrows():
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font("Helvetica", style="B", size=14)
            pdf.cell(0, 10, "Validation Table by Sample (continued)", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="L")
            pdf.ln(5)
            _draw_table_header(pdf, TABLE_COLS, TABLE_WIDTHS)
            pdf.set_font("Helvetica", size=8)

        pdf.set_fill_color(255, 255, 255)
        for col, w in zip(TABLE_COLS, TABLE_WIDTHS):
            value = str(row[col]) if (col in row and not pd.isna(row[col])) else "NA"
            key   = value.upper().strip()
            r, g, b = FILL_COLORS.get(key, (255, 255, 255))
            pdf.set_fill_color(r, g, b)
            pdf.cell(w, 8, "OK" if key == "OK" else value, border=1, fill=True)
        pdf.ln()

    pdf.output(output_path)
    print(f"PDF report saved: {output_path}")


# ============================================================
# EXCEL EXPORT
# ============================================================

# Optional hydrological fields are not calculated by this component. They are
# exported only when they already exist in the validated CSV data, allowing a
# future template to pass them through to All_Validated_Data.xlsx.
OPTIONAL_PASSTHROUGH_COLUMNS = ["q", "hg", "f", "cnr", "sio2", "ALL"]

EXPORT_COLUMNS = [
    "SampleID", "SiteCode", "SiteName", "Type", "year", "month",
    "StartDate", "EndDate",
    "CL(mg/l)", "SO4S(mg/l)", "NO3N(mg/l)", "PO4P(mg/l)",
    "CA(mg/l)", "MG(mg/l)", "NA(mg/l)", "K(mg/l)",
    "AL(mg/l)", "FE(mg/l)", "MN(mg/l)",
    "AS(mg/l)", "CD(mg/l)", "CR(mg/l)", "CU(mg/l)", "CO(mg/l)",
    "MO(mg/l)", "NI(mg/l)", "PB(mg/l)", "ZN(mg/l)",
    "P(mg/l)", "S(mg/l)",
    "NH4N(mg/l)", "TN(mg/l)", "DOC(mg/l)",
    "NING(mg/l)", "NDON(mg/l)",
    "WeightedpH", "WeightedConductivity(µS/cm)", "Temperature(ºC)",
    "Volume(ml)", "Precip(l/m2)", "AlkalinityICPForests(µeq/l)",
    *OPTIONAL_PASSTHROUGH_COLUMNS,
    "Metals_SW(µeq/l)", "Org-(µeq/l)", "SumAnions(µeq/l)",
    "+Org(µeq/l)", "H(µeq/l)", "SumCations(µeq/l)",
    "sC - sA IonsDiff.%", "sC - sA QualityIonsBalance",
    "sC - sA - Org- IonsDiff.%", "IonsDiffOrg.Limit(%)",
    "IonsDiffOrg.OverLimit.pp", "IonsDiffOrg.OverLimit.relative%",
    "RatioNa/Cl", "ConductivityCalculatedWithoutCorrection(µS/cm)",
    "IonicStrenght(mol/l)", "IonicActivityFactor",
    "ConductivityCalculatedCorrected(µS/cm)", "Cond. Diff.%Cc-Xm",
    "Ions balance Org", "Ratio Na/Cl", "Conductivity", "OrgN", "VAL",
]



def export_excel(df, sampling_ty, output_path, filter_val=None):
    """
    Exports a DataFrame to Excel, optionally filtering by VAL.
    Sorts rows by year / month / custom SampleID order from samplesInfo.xlsx.
    Highlights cells with missing values in red.

    filter_val: if "NO", keeps only rows where VAL == "NO".
                if None, keeps all rows.
    """
    d = prepare_report_data(df)

    if filter_val == "NO":
        d = d[d["VAL"] == "NO"].copy()

    # Remove empty SampleID rows
    d = d[
        d["SampleID"].notna() &
        (d["SampleID"].astype(str).str.strip() != "") &
        (d["SampleID"].astype(str).str.upper().str.strip() != "NA")
    ].copy()

    d["year"]  = pd.to_numeric(d["year"],  errors="coerce")
    d["month"] = pd.to_numeric(d["month"], errors="coerce")

    # Build custom sort order from samplesInfo
    order_ids = [
        clean_id(x) for x in sampling_ty["SampleID"].tolist()
        if pd.notna(x) and str(x).strip() != ""
    ]
    combined_order = []
    for x in order_ids:
        combined_order.append(x)
        combined_order.append(x + "REP")

    d["_clean"] = d["SampleID"].apply(clean_id)
    d["_order"] = pd.Categorical(d["_clean"], categories=combined_order, ordered=True)
    d = d.sort_values(by=["year", "month", "_order"]).drop(columns=["_clean", "_order"])
    d["SampleID"] = d["SampleID"].apply(clean_id)

    # Keep only defined export columns that are present
    cols = [c for c in EXPORT_COLUMNS if c in d.columns]
    d[cols].to_excel(output_path, index=False)

    # Highlight empty cells in red
    wb = load_workbook(output_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None or cell.value == "":
                cell.fill = red_fill
    wb.save(output_path)
    print(f"Excel saved: {output_path}")


# ============================================================
# MAIN: load data, generate outputs
# ============================================================

# Merge all validated CSV files into a single DataFrame
csv_files = [f for f in os.listdir(extract_dir) if f.endswith(".csv")]
if not csv_files:
    raise RuntimeError(f"No CSV files found in {extract_dir}")

df_final = pd.concat(
    [pd.read_csv(os.path.join(extract_dir, f), sep="\t") for f in csv_files],
    ignore_index=True,
)
print(f"Loaded {len(df_final)} rows from {len(csv_files)} CSV files")

# Fill typology for REP samples
df_final = add_typology_from_sampling_info(df_final, sampling_ty)

# Sort and deduplicate: keep row with fewest NaNs per SampleID/month
df_final["_num_nans"] = df_final.isna().sum(axis=1)
df_final = (
    df_final
    .sort_values(by=["SampleID", "month", "_num_nans"])
    .drop_duplicates(subset=["SampleID", "month"], keep="first")
    .drop(columns="_num_nans")
    .reset_index(drop=True)
)

# Remove empty SampleID rows
df_final = df_final[
    df_final["SampleID"].notna() &
    (df_final["SampleID"].astype(str).str.strip() != "") &
    (df_final["SampleID"].astype(str).str.upper().str.strip() != "NA")
].reset_index(drop=True)

print(f"Final dataset: {len(df_final)} rows after deduplication")

# Generate charts
chart_paths = create_validation_charts(df_final, charts_dir)

# Generate PDF
generate_pdf_report(df_final, sampling_ty, chart_paths, output_pdf_path)

# Generate Excel outputs
export_excel(df_final, sampling_ty, output_repeat_path, filter_val="NO")
export_excel(df_final, sampling_ty, output_all_path,    filter_val=None)

print("--- ALL OUTPUTS GENERATED ---")